import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font

from create import Create
from tabulate import tabulate
from model import Model
from process import Process
import pandas as pd

#
# c = Create('CREATOR', 20, 'erlang', 2)
#
# p1 = Process('SPEAKER-1', 2.5, 'uniform', 3.5, 9, 5)
# p2 = Process('SPEAKER-2', 2.5, 'uniform', 3.5, 9, 5)
# p3 = Process('SPEAKER-3', 2.5, 'uniform', 3.5, 9, 5)
#
# c.next_element = [p1, p2, p3]
# c.probability = [1 / 3, 1 / 3, 1 / 3]
#
# # p1.next_element = [p2, p3]
# # p1.probability = [0.5, 0.5]
# # p2.next_element = [p3, p1]
# # p2.probability = [0.5, 0.5]
# # p3.next_element = [p2, p1]
# # p3.probability = [0.5, 0.5]
#
# # p1.max_queue = 5
# # p2.max_queue = 5
# # p3.max_queue = 5
#
# # elements = [c, p1]
# elements = [c, p1, p2, p3]
# model = Model(elements)
# res = model.simulate(16 * 60)

workbook = openpyxl.Workbook()
sheet = workbook.active

min_payback_period = float('inf')
min_payback_period_row = 0
tests = 10
max_pause_count = [3, 6, 8, 10, 13, 15, 16, 17, 18, 20]
pause_duration = [2.5, 3, 3.5, 4, 2.5, 3, 3.5, 4, 2.5, 3]
column_names = [
    'quantity_create',
    'max_pause_count',
    'pause_duration',
    'quantity_p1',
    'fail1',
    'queue1',
    'remove1',
    'count_discount1',
    'quantity_p2',
    'fail2',
    'queue2',
    'remove2',
    'count_discount2',
    'quantity_p3',
    'fail3',
    'queue3',
    'remove3',
    'count_discount3',
    'Money for day',
    'Payback period in years'
]

for i in range(tests):
    c = Create('CREATOR', 20, 'erlang', 2)

    p1 = Process('SPEAKER-1', 2.5, 'uniform', 3.5, max_pause_count[i], 5)
    p2 = Process('SPEAKER-2', 2.5, 'uniform', 3.5, max_pause_count[i], 5)
    p3 = Process('SPEAKER-3', 2.5, 'uniform', 3.5, max_pause_count[i], 5)

    c.next_element = [p1, p2, p3]
    c.probability = [1 / 3, 1 / 3, 1 / 3]
    p1.pause_duration = pause_duration[i]  # 3
    p2.pause_duration = pause_duration[i]  # 3
    p3.pause_duration = pause_duration[i]  # 3
    # p1.next_element = [p2, p3]
    # p1.probability = [0.5, 0.5]
    # p2.next_element = [p3, p1]
    # p2.probability = [0.5, 0.5]
    # p3.next_element = [p2, p1]
    # p3.probability = [0.5, 0.5]

    # p1.max_queue = 5
    # p2.max_queue = 5
    # p3.max_queue = 5

    # elements = [c, p1]
    elements = [c, p1, p2, p3]
    model = Model(elements)
    res = model.simulate(16 * 60)

    if res[-1][4] < min_payback_period:
        min_payback_period = res[-1][4]
        min_payback_period_row = i + 2

    result1 = {
        'quantity_create': res[0],
        'max_pause_count': max_pause_count[i],
        'pause_duration': res[-1][0],
        'quantity_p1': res[1],
        'fail1': res[2][1],
        'queue1': res[2][0],
        'remove1': res[2][2],
        'count_discount1': res[2][3],
        'quantity_p2': res[3],
        'fail2': res[4][1],
        'queue2': res[4][0],
        'remove2': res[4][2],
        'count_discount2': res[4][3],
        'quantity_p3': res[5],
        'fail3': res[6][1],
        'queue3': res[6][0],
        'remove3': res[6][2],
        'count_discount3': res[6][3],
        'Money for day': res[-1][1],
        'Payback period in years': res[-1][4]
    }

    if i == 0:
        for col_num, column_name in enumerate(column_names, start=1):
            sheet.cell(row=1, column=col_num, value=column_name)

    for col_num, value in enumerate(result1.values(), start=1):
        sheet.cell(row=i + 2, column=col_num, value=value)

for cell in sheet[min_payback_period_row]:
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for cell in sheet[1]:
    cell.font = Font(bold=True)

workbook.save("RESULTS.xlsx")
